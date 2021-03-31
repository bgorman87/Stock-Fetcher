import json
# import re
import time
# import random
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup

print_bool = False


def get_roe_npv(discount_rate, equity, equity_rate, shares_outstanding, dividend_rate, growth_rate, margin_of_safety):
    conservative_growth = growth_rate * (1 - margin_of_safety)
    shareholders_equity = []
    dividend = []
    npv_dividend = []
    shareholders_equity.append(equity * (1 + conservative_growth) / shares_outstanding)
    dividend.append(dividend_rate * (1 + conservative_growth))
    npv_dividend.append(dividend[0])
    for i in range(2, 11):
        shareholders_equity.append(shareholders_equity[i - 2] * (1 + conservative_growth))
        dividend.append(dividend[i - 2] * (1 + conservative_growth))
        npv_dividend.append(dividend[i - 1] / ((1 + discount_rate) ** (i - 1)))
    y10_net_income = shareholders_equity[-1] * equity_rate
    required_value = y10_net_income / discount_rate
    npv_required_value = required_value / ((1 + discount_rate) ** len(dividend))
    npv_dividends = sum(npv_dividend)
    npv_value = npv_required_value + npv_dividends
    return int(npv_value)


def get_dcf_npv(discount_rate, cash_and_cash_eq, liabilities, free_cash_flow, outstanding_shares,
                growth_rate, margin_of_safety):
    conservative_growth = growth_rate * (1 - margin_of_safety)
    growth_decline = 0.05
    free_cash_growth = []
    npv_free_cash = []
    free_cash_growth.append(free_cash_flow * (1 + conservative_growth))
    npv_free_cash.append(free_cash_growth[0] / (1 + discount_rate))
    for i in range(2, 11):
        free_cash_growth.append(
            free_cash_growth[i - 2] * (1 + (conservative_growth * ((1 - growth_decline) ** (i - 1)))))
        npv_free_cash.append(free_cash_growth[i - 1] / ((1 + discount_rate) ** i))

    total_npv = sum(npv_free_cash)
    year_10_free_cash = npv_free_cash[-1] * 12

    npv_dcf = (total_npv + year_10_free_cash + cash_and_cash_eq - liabilities) / outstanding_shares

    return npv_dcf


def row_get_Data_Text(tr, coltag='td'):  # td (data) or th (header)
    return [td.get_text(strip=True) for td in tr.find_all(coltag)]


def get_morningstar_pe(local_symbol):
    historical_PE_link = "http://financials.morningstar.com/valuate/current-valuation-list.action?&t" \
                         "={}&region=can&culture=en-US"
    link = historical_PE_link.format(local_symbol)
    values = []
    try:
        html_content = requests.get(link).text
        soup = BeautifulSoup(html_content, "lxml")
        table = soup.find(
            lambda tag: tag.name == 'table' and tag.has_attr('id') and tag['id'] == "currentValuationTable")
        rows = table.findAll(lambda tag: tag.name == 'tr')
        for row in rows:
            values.append(row_get_Data_Text(row, 'td'))
    except (requests.exceptions.RequestException, ValueError, AttributeError, KeyError, TypeError):
        pass
        # print(e)
    if values:
        return values[3][3]
    else:
        return values


def get_morningstar_roe(local_symbol):
    historical_ROE_link = "http://financials.morningstar.com/finan/financials/getKeyStatPart.html?&t={}"
    link = historical_ROE_link.format(local_symbol)
    values = []
    content_rows = []
    try:
        html_content = requests.get(link).text
        test = json.loads(html_content)
        soup = BeautifulSoup(test['componentData'], "html5lib")
        rows = soup.findAll(lambda tag: tag.name == 'td')

        for row in rows:
            if "i26" in str(row):
                content_rows.append(row)
    except (requests.exceptions.RequestException, ValueError, AttributeError, KeyError, TypeError):
        pass
        # print(e)
    if content_rows:
        values = [value.get_text() for value in content_rows]
    for i, value in enumerate(values):
        if not is_float(value):
            values[i] = 0
    if values:
        return sum([float(value) for value in values[5:-1]]) / 5
    else:
        return values


# def get_yahoo_stat(local_symbol, stat_type=None):
#     rand_int = random.randint(4, 6)
#     time.sleep(rand_int)
#
#     current_eps_link = "https://ca.finance.yahoo.com/quote/{}?p={}"
#     growth_estimate_link = "https://ca.finance.yahoo.com/quote/{}/analysis?p={}"
#     fcf_link = "https://ca.finance.yahoo.com/quote/{}/cash-flow?p={}"
#     ceq_link = "https://ca.finance.yahoo.com/quote/{}/balance-sheet?p={}"
#     shares_outstanding_link = "https://ca.finance.yahoo.com/quote/{}/key-statistics?p={}"
#     profile_link = "https://ca.finance.yahoo.com/quote/{}/profile?p={}"
#     sustainability_link = "https://ca.finance.yahoo.com/quote/{}/sustainability?p={}"
#     if stat_type == "eps":
#         link = current_eps_link.format(local_symbol, local_symbol)
#     elif stat_type == "growth" or stat_type == "title":
#         link = growth_estimate_link.format(local_symbol, local_symbol)
#     elif stat_type == "fcf":
#         link = fcf_link.format(local_symbol, local_symbol)
#     elif stat_type == "ceq":
#         link = ceq_link.format(local_symbol, local_symbol)
#     elif stat_type == "shares_outstanding":
#         link = shares_outstanding_link.format(local_symbol, local_symbol)
#     elif stat_type == "profile":
#         link = profile_link.format(local_symbol, local_symbol)
#     elif stat_type == "sustainability":
#         link = sustainability_link.format(local_symbol, local_symbol)
#     else:
#         link = ""
#
#     data = ""
#     error = False
#     json_data = {}
#     html_content = ""
#
#     try:
#         html_content = requests.get(link, timeout=None).text
#     except requests.exceptions.RequestException as reqE:
#         print(reqE)
#         error = True
#     if not error:
#         try:
#             soup = BeautifulSoup(html_content, 'lxml')
#             pattern = re.compile(r'\s--\sData\s--\s')
#             script_data = soup.find('script', text=pattern).contents[0]
#             start = script_data.find("context") - 2
#             json_data = json.loads(script_data[start:-12])
#         except (AttributeError, KeyError, ValueError) as soupError:
#             if print_bool:
#                 print(soupError)
#             error = True
#     if not error and stat_type == "eps":
#         try:
#             eps = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['defaultKeyStatistics'] \
#                 ['trailingEps']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             eps = ""
#         try:
#             price = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['price'] \
#                 ['regularMarketPrice']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             price = ""
#         data = [eps, price]
#     elif stat_type == "growth":
#         try:
#             growth = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['earningsTrend']['trend'][4][
#                 'growth']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             growth = ""
#         data = growth
#     elif stat_type == "profile":
#         try:
#             industry = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['assetProfile']['industry']
#         except (KeyError, IndexError, TypeError, ValueError):
#             industry = ""
#         try:
#             summary = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['assetProfile'][
#             'longBusinessSummary']
#         except (KeyError, IndexError, TypeError, ValueError):
#             summary = ""
#         data = [industry, summary]
#     elif stat_type == "sustainability":
#         try:
#             esg_score = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['esgScores']["totalEsg"][
#                 "raw"]
#         except (KeyError, IndexError, TypeError, ValueError):
#             esg_score = ""
#         try:
#             controversy_level = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['esgScores'][
#                 'highestControversy']
#         except (KeyError, IndexError, TypeError, ValueError):
#             controversy_level = ""
#         data = [esg_score, controversy_level]
#     elif stat_type == "title":
#         try:
#             local_title = json_data['context']['dispatcher']['stores']['PageStore']['pageData']['title']
#         except (KeyError, IndexError, TypeError, ValueError):
#             local_title = ""
#         data = local_title
#     elif stat_type == "fcf":
#         try:
#          trailing_cash_flow_raw = json_data['context']['dispatcher']['stores']['QuoteTimeSeriesStore']['timeSeries'][
#                 'trailingFreeCashFlow'][0]['reportedValue']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             trailing_cash_flow_raw = ""
#         try:
#          trailing_cash_flow_fmt = json_data['context']['dispatcher']['stores']['QuoteTimeSeriesStore']['timeSeries'][
#                 'trailingFreeCashFlow'][0]['reportedValue']['fmt']
#         except (KeyError, IndexError, TypeError, ValueError):
#             trailing_cash_flow_fmt = ""
#         data = [trailing_cash_flow_raw, trailing_cash_flow_fmt]
#     elif stat_type == "ceq":
#         try:
#             cash_raw = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistory'][
#                 'balanceSheetStatements'][0]['cash']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             cash_raw = ""
#         try:
#             cash_fmt = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistory'][
#                 'balanceSheetStatements'][0]['cash']['fmt']
#         except (KeyError, IndexError, TypeError, ValueError):
#             cash_fmt = ""
#         try:
#             total_liabilities_raw = json_data['context']['dispatcher']['stores']['QuoteSummaryStore'][
#                 'balanceSheetHistory']['balanceSheetStatements'][0]['totalLiab']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_liabilities_raw = ""
#         try:
#             total_liabilities_fmt = json_data['context']['dispatcher']['stores']['QuoteSummaryStore'][
#                 'balanceSheetHistory']['balanceSheetStatements'][0]['totalLiab']['fmt']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_liabilities_fmt = ""
#         try:
#             total_stock_equity_raw = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistory'][
#                     'balanceSheetStatements'][0]['totalStockholderEquity']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_stock_equity_raw = ""
#         try:
#             total_stock_equity_fmt = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistory'][
#                     'balanceSheetStatements'][0]['totalStockholderEquity']['fmt']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_stock_equity_fmt = ""
#         try:
#             total_quarterly_assets_raw = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistoryQuarterly'][
#                     'balanceSheetStatements'][1]['totalAssets']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_quarterly_assets_raw = ""
#         try:
#             total_quarterly_liabilities_raw = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistoryQuarterly'][
#                     'balanceSheetStatements'][1]['totalLiab']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_quarterly_liabilities_raw = ""
#         try:
#             total_long_term_debt = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['balanceSheetHistoryQuarterly'][
#                     'balanceSheetStatements'][1]['longTermDebt']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             total_long_term_debt = ""
#         try:
#             net_income = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['incomeStatementHistory'][
#                     'incomeStatementHistory'][0]['netIncome']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             net_income = ""
#         try:
#             revenue = \
#                 json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['incomeStatementHistory'][
#                     'incomeStatementHistory'][0]['totalRevenue']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             revenue = ""
#         data = [cash_raw, cash_fmt, total_liabilities_raw, total_liabilities_fmt, total_stock_equity_raw,
#                 total_stock_equity_fmt, total_quarterly_assets_raw, total_quarterly_liabilities_raw,
#                 total_long_term_debt, net_income, revenue]
#     elif stat_type == "shares_outstanding":
#         try:
#            shares_out_raw = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['defaultKeyStatistics'][
#                 'sharesOutstanding']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             shares_out_raw = ""
#         try:
#            shares_out_fmt = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['defaultKeyStatistics'][
#                 'sharesOutstanding']['fmt']
#         except (KeyError, IndexError, TypeError, ValueError):
#             shares_out_fmt = ""
#         try:
#            trailing_dividend_raw = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['summaryDetail'][
#                 'trailingAnnualDividendRate']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             trailing_dividend_raw = ""
#         try:
#            trailing_dividend_fmt = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['summaryDetail'][
#                 'trailingAnnualDividendRate']['fmt']
#         except (KeyError, IndexError, TypeError, ValueError):
#             trailing_dividend_fmt = ""
#         try:
#             market_cap = json_data['context']['dispatcher']['stores']['QuoteSummaryStore']['summaryDetail'][
#                 'marketCap']['raw']
#         except (KeyError, IndexError, TypeError, ValueError):
#             market_cap = ""
#         data = [shares_out_raw, shares_out_fmt, trailing_dividend_raw, trailing_dividend_fmt, market_cap]
#     return data


def is_float(f):
    try:
        float(f)
        return True
    except (TypeError, ValueError):
        return False


def update_spreadsheet(symbol_list):
    template_file = "Stonks Files\\Overall Good Symbol Spreadsheets\\Overall Good Symbols Template.xlsx"
    file_save = "Stonks Files\\Overall Good Symbol Spreadsheets\\Stonks Data - " + \
                time.strftime("%Y-%b-%d -- %I %M %p") + ".xlsx"
    with open('Stonks Files\\Output\\stock_data.txt') as json_file:
        stonks = json.load(json_file)
    wb = load_workbook(template_file)
    stonks_ws = wb.active
    row = 2
    for symbol in symbol_list:
        try:
            _ = stonks[symbol]
            stonks_ws.cell(row, 1).value = symbol
            stonks_ws.cell(row, 2).value = stonks[symbol]["Title"]
            stonks_ws.cell(row, 3).value = stonks[symbol]["Quality"]
            stonks_ws.cell(row, 4).value = stonks[symbol]["Details"]["Industry"]
            stonks_ws.cell(row, 5).value = stonks[symbol]["Current"]
            stonks_ws.cell(row, 6).value = stonks[symbol]["P/E"]
            stonks_ws.cell(row, 7).value = stonks[symbol]["DCF"]
            stonks_ws.cell(row, 8).value = stonks[symbol]["ROE"]
            try:
                _ = stonks[symbol]["Details"]
                stonks_ws.cell(row, 10).value = stonks[symbol]["Details"]["Market Cap"]
                stonks_ws.cell(row, 11).value = stonks[symbol]["Details"]["Revenue"]
                stonks_ws.cell(row, 12).value = stonks[symbol]["Details"]["Net Income"]
                stonks_ws.cell(row, 13).value = stonks[symbol]["Details"]["Assets"]
                stonks_ws.cell(row, 14).value = stonks[symbol]["Details"]["Liabilities"]
                stonks_ws.cell(row, 16).value = stonks[symbol]["Details"]["Debt"]
                stonks_ws.cell(row, 18).value = stonks[symbol]["Details"]["ESG Score"]
                stonks_ws.cell(row, 19).value = stonks[symbol]["Details"]["Controversy"]
                stonks_ws.cell(row, 20).value = stonks[symbol]["Details"]["Summary"]
            except KeyError:
                continue
            row += 1
        except KeyError:
            continue
# current_prices, current_ratios, symbols = lowest_price(good_symbols=good_symbols)
    # zipped = list(zip(current_ratios, current_prices, symbols))
    # zipped.sort(reverse=True)
    # print("\n", zipped)
    wb.save(file_save)
    wb.close()
