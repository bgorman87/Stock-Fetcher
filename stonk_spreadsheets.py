import logging
import re
import time
from typing import List

import yahooquery
from openpyxl import load_workbook

from stonks import (
    JustSkip,
    RecentlyUpdated,
    calculate_and_update_dcf_value,
    calculate_and_update_pe_value,
    calculate_and_update_roe_value,
    fetch_stock_data,
    is_float,
)
from stonk_list import DB_FILE_PATH, connect_to_database
from utils import Stock

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# Constants
TEMPLATE_FILE = "Stonks Files/Individual Good Symbol Spreadsheets/!Individual Good Symbol Template.xlsx"
OVERALL_TEMPLATE_FILE = (
    "Stonks Files/Overall Good Symbol Spreadsheets/Overall Good Symbols Template.xlsx"
)
GOOD_SYMBOLS_DIRECTORY = "Stonks Files/Individual Good Symbol Spreadsheets/"
OVERALL_SYMBOLS_DIRECTORY = "Stonks Files/Overall Good Symbol Spreadsheets/"
EXCEL_UPDATE_INTERVAL = 86400 / 2  # 12 hours in seconds


def update_good_spreadsheets(good_stocks: List[Stock]):
    logging.info("Starting update for individual 'good' spreadsheets")
    good_symbols_to_update = get_symbols_to_update(good_stocks)
    yahoo_symbols = format_symbols_for_yahoo(good_symbols_to_update)
    stock_data = fetch_stock_data_from_yahoo(yahoo_symbols)

    total_symbols = len(good_symbols_to_update)
    for idx, symbol in enumerate(good_symbols_to_update):
        try:
            logging.info(
                f"Updating spreadsheet for {symbol} ({idx + 1}/{total_symbols})"
            )
            update_individual_spreadsheet(
                symbol, yahoo_symbols[idx], stock_data[symbol]
            )
        except RecentlyUpdated:
            logging.info(f"{symbol} has been recently updated. Skipping update.")
            continue
        except Exception as e:
            logging.error(f"Error updating spreadsheet for {symbol}: {e}")
            continue


def update_spreadsheet(symbol_list: List[str]):
    logging.info("Starting update for the overall 'good' spreadsheet")
    file_save = f"{OVERALL_SYMBOLS_DIRECTORY}Stonks Data - {time.strftime('%Y-%b-%d -- %I %M %p')}.xlsx"

    with load_workbook(OVERALL_TEMPLATE_FILE) as wb:
        stonks_ws = wb.active
        for idx, symbol in enumerate(symbol_list):
            logging.info(
                f"Updating overall spreadsheet for {symbol} ({idx + 1}/{len(symbol_list)})"
            )
            try:
                update_overall_spreadsheet_row(stonks_ws, symbol, idx + 2)
            except KeyError:
                logging.warning(f"Key error while updating {symbol}. Skipping.")
                continue
        wb.save(file_save)
    logging.info(f"Overall 'good' spreadsheet updated and saved to {file_save}")


def get_symbols_to_update(good_symbols: List[str]) -> List[str]:
    """Retrieve symbols that need to be updated."""
    logging.info("Checking which symbols need an update")
    symbols_to_update = []
    for symbol in good_symbols:
        if should_update_symbol(symbol):
            symbols_to_update.append(symbol)
    return symbols_to_update


def should_update_symbol(symbol: str) -> bool:
    """Check if the symbol needs to be updated based on the last update time."""
    try:
        with connect_to_database(DB_FILE_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT last_updated FROM stonks WHERE symbol=?", (symbol,))
            last_updated = cur.fetchone()
            if last_updated and time.time() - last_updated[0] > EXCEL_UPDATE_INTERVAL:
                logging.debug(f"Symbol {symbol} needs an update.")
                return True
    except (KeyError, ValueError):
        logging.error(f"Error checking update status for symbol {symbol}")
    return False


def fetch_stock_data_from_yahoo(symbols: List[str]) -> dict:
    """Fetch stock data from Yahoo."""
    logging.info("Fetching stock data from Yahoo Finance")
    tickers = yahooquery.Ticker(symbols)
    return tickers.all_modules


def update_individual_spreadsheet(symbol: str, yahoo_symbol: str, stock_data: dict):
    """Update individual spreadsheets with stock data."""
    file_title = sanitize_filename(get_symbol_title(symbol))
    file_save = f"{GOOD_SYMBOLS_DIRECTORY}{symbol.split('.')[0]} - {file_title}.xlsx"

    if not is_recently_updated(file_save):
        logging.info(f"Stock details for {symbol} need an update")
        stock_details = get_stock_details(symbol, yahoo_symbol, stock_data)
        save_stock_data_to_spreadsheet(file_save, stock_details, stock_data, symbol)
    else:
        logging.info(f"Spreadsheet for {symbol} is already up-to-date")


def get_symbol_title(symbol: str) -> str:
    """Retrieve the title for the symbol."""
    try:
        with connect_to_database(DB_FILE_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT title FROM stonks WHERE symbol=?", (symbol,))
            title = cur.fetchone()
        return title[0] if title else symbol
    except Exception as e:
        logging.error(f"Error fetching title for symbol {symbol}: {e}")
        return symbol


def sanitize_filename(filename: str) -> str:
    """Sanitize the filename to remove invalid characters."""
    return re.sub(r"(?u)[^-\w.]", "", filename.strip().replace(" ", "_"))


def is_recently_updated(file_path: str) -> bool:
    """Check if the file was recently updated."""
    try:
        with load_workbook(file_path) as wb:
            updated_time = wb["Stats"].cell(1, 5).value
            if time.time() - int(updated_time) < EXCEL_UPDATE_INTERVAL:
                logging.info(f"File {file_path} is already up-to-date.")
                raise RecentlyUpdated
    except (FileNotFoundError, TypeError, KeyError):
        logging.debug(f"File {file_path} needs an update.")
        return False
    return True


def get_stock_details(symbol: str, yahoo_symbol: str, stock_data: dict):
    """Retrieve stock details either from the database or Yahoo."""
    try:
        with connect_to_database(DB_FILE_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM stonks WHERE symbol=?", (symbol,))
            details = cur.fetchone()
            if details:
                logging.info(f"Fetched stock details for {symbol} from the database")
                return details
            else:
                logging.info(
                    f"Fetching and calculating stock details for {symbol} from Yahoo"
                )
                return fetch_and_calculate_stock_details(yahoo_symbol)
    except (TypeError, ValueError, KeyError) as e:
        logging.error(f"Error fetching stock details for {symbol}: {e}")
        return fetch_and_calculate_stock_details(yahoo_symbol)


def fetch_and_calculate_stock_details(symbol: str):
    """Fetch and calculate stock details from Yahoo."""
    try:
        details = fetch_stock_data(symbol)
        if not is_float(details.get("growthEstimate")):
            logging.warning(f"No valid growth estimate for {symbol}. Skipping.")
            raise JustSkip(f"No growth estimate for {symbol}")
        return calculate_stock_values(details)
    except JustSkip as e:
        logging.warning(e)
        return []


def calculate_stock_values(stock_data: dict):
    """Calculate stock values based on fetched data."""
    logging.info("Calculating stock values based on fetched data")
    try:
        discount_rate = 0.09
        current_5y_backtrack_pe = calculate_and_update_pe_value(
            stock_data.get("currentEPS"),
            stock_data.get("historicalPE"),
            stock_data.get("growthEstimate"),
            discount_rate,
        )
        current_10y_backtrack_dcf = calculate_and_update_dcf_value(
            stock_data.get("cash"),
            stock_data.get("liabilities"),
            stock_data.get("freeCashFlow"),
            stock_data.get("sharesOutstanding"),
            stock_data.get("growthEstimate"),
        )
        current_10y_backtrack_roe = calculate_and_update_roe_value(
            stock_data.get("stockholdersEquity"),
            stock_data.get("historicalROE"),
            stock_data.get("sharesOutstanding"),
            stock_data.get("trailingDividendRate"),
            stock_data.get("growthEstimate"),
        )
        return [
            stock_data.get("title"),
            stock_data.get("industry"),
            stock_data.get("currentPrice"),
            stock_data.get("quarterlyLiabilities"),
            stock_data.get("quarterlyAssets"),
            stock_data.get("longTermDebt"),
            stock_data.get("netIncome"),
            stock_data.get("revenue"),
            stock_data.get("marketCap"),
            stock_data.get("growthEstimate"),
            current_5y_backtrack_pe,
            current_10y_backtrack_dcf,
            current_10y_backtrack_roe,
        ]
    except Exception as e:
        logging.error(f"Error calculating stock values: {e}")
        return []


def save_stock_data_to_spreadsheet(
    file_save: str, stock_details: list, stock_data: dict, symbol: str
):
    """Save stock data to a spreadsheet."""
    try:
        with load_workbook(TEMPLATE_FILE) as wb:
            ws = wb["Stats"]
            ws.cell(1, 5).value = time.time()
            write_stock_details_to_sheet(ws, stock_details, symbol)
            wb.save(file_save)
        logging.info(f"Spreadsheet for {symbol} saved to {file_save}")
    except Exception as e:
        logging.error(f"Error saving spreadsheet for {symbol}: {e}")


def write_stock_details_to_sheet(ws, stock_details: list, symbol: str):
    """Write stock details to the worksheet."""
    ws.cell(3, 1).value = symbol
    for idx, value in enumerate(stock_details):
        ws.cell(3, idx + 2).value = value


def update_overall_spreadsheet_row(ws, symbol: str, row: int):
    """Update a row in the overall spreadsheet with stock data."""
    try:
        with connect_to_database(DB_FILE_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM stonks WHERE symbol=?", (symbol,))
            stonk_details = cur.fetchone()

        if stonk_details:
            file_title = sanitize_filename(stonk_details[8])
            spreadsheet_link = (
                f"{GOOD_SYMBOLS_DIRECTORY}{symbol.split('.')[0]} - {file_title}.xlsx"
            )
            ws.cell(row, 1).hyperlink = spreadsheet_link
            ws.cell(row, 1).value = symbol
            ws.cell(row, 2).value = stonk_details[8]
            ws.cell(row, 3).value = stonk_details[7]
            ws.cell(row, 4).value = stonk_details[9]
            ws.cell(row, 5).value = stonk_details[2]
            ws.cell(row, 6).value = stonk_details[3]
            ws.cell(row, 7).value = stonk_details[4]
            ws.cell(row, 8).value = stonk_details[5]
            ws.cell(row, 10).value = stonk_details[10]
            ws.cell(row, 11).value = stonk_details[11]
            ws.cell(row, 12).value = stonk_details[12]
            ws.cell(row, 13).value = stonk_details[13]
            ws.cell(row, 14).value = stonk_details[14]
            ws.cell(row, 16).value = stonk_details[15]
            ws.cell(row, 18).value = stonk_details[16]
            ws.cell(row, 19).value = stonk_details[17]
            ws.cell(row, 20).value = stonk_details[18]
        logging.info(f"Row for {symbol} updated in overall spreadsheet")
    except Exception as e:
        logging.error(f"Error updating row for {symbol} in overall spreadsheet: {e}")
